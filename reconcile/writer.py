"""
writer.py
Builds the final, formatted Excel workbook finance will actually open:
  - Summary        : headline counts and totals control-check
  - Full_Matches   : matches found by deterministic exact matching
                      (reference match, or amount+date match)
  - Fuzzy_Matches  : matches found by fuzzy description matching or the
                      LLM disambiguation step -- kept separate from
                      Full_Matches since these carry more residual risk
                      and are worth a reviewer's separate glance
  - Needs_Review   : rows the LLM looked at but did NOT auto-accept
                      (low confidence, or no plausible candidate) --
                      genuinely needs a human decision
  - Bank_Only      : bank rows with no ledger counterpart at all
  - Ledger_Only    : ledger rows with no bank counterpart at all
  - LLM_Audit_Log  : every LLM call made, for traceability
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
GOOD_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
WARN_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

# Which match "method" values belong on which output sheet.
EXACT_METHODS = {"exact_reference", "exact_amount_date"}
FUZZY_METHODS = {"fuzzy_auto", "llm_match"}

# Column order for the two match sheets -- both-sides columns first (the
# most useful thing for a reviewer to see side by side), then the
# legacy shared date/amount/description (kept for backward compatibility
# with anything still reading those keys), then method/confidence.
MATCH_COLS = [
    "bank_row", "ledger_row",
    "bank_date", "ledger_date",
    "bank_amount", "ledger_amount",
    "bank_description", "ledger_description",
    "date", "amount", "description",
    "method", "confidence",
]


def _write_df_sheet(wb, sheet_name, df: pd.DataFrame, row_fill=None):
    ws = wb.create_sheet(sheet_name)
    if df.empty:
        ws.append(["No rows"])
        return ws

    ws.append(list(df.columns))
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    for _, row in df.iterrows():
        values = [
            v.strftime("%Y-%m-%d") if isinstance(v, pd.Timestamp) else v
            for v in row.tolist()
        ]
        ws.append(values)
        if row_fill:
            for cell in ws[ws.max_row]:
                cell.fill = row_fill

    for i, col in enumerate(df.columns, start=1):
        col_len = df[col].apply(lambda x: len(str(x))).max() if len(df) else 0
        max_len = max(int(col_len), len(col)) + 2
        ws.column_dimensions[get_column_letter(i)].width = min(max_len, 45)

    ws.freeze_panes = "A2"
    return ws


def _prepare_match_sheet_df(matches: list[dict]) -> pd.DataFrame:
    """Builds a DataFrame with a consistent column set for a match sheet,
    filling in any missing both-sides columns with None so older match
    dicts (e.g. from a caller that hasn't been updated) don't break the
    export -- they just show blank in the new columns instead."""
    df = pd.DataFrame(matches)
    if df.empty:
        return df
    for c in MATCH_COLS:
        if c not in df.columns:
            df[c] = None
    ordered = MATCH_COLS + [c for c in df.columns if c not in MATCH_COLS]
    return df[ordered]


def write_output(
    path: str,
    matches: list[dict],
    bank: pd.DataFrame,
    ledger: pd.DataFrame,
    llm_log: list,
    llm_confidence_threshold: float = 0.85,
):
    wb = Workbook()
    wb.remove(wb.active)

    exact_matches = [m for m in matches if m.get("method") in EXACT_METHODS]
    fuzzy_matches = [m for m in matches if m.get("method") in FUZZY_METHODS]
    # Anything with an unrecognized/missing method still gets exported
    # somewhere rather than silently vanishing from the workbook.
    other_matches = [m for m in matches if m.get("method") not in EXACT_METHODS | FUZZY_METHODS]
    fuzzy_matches += other_matches

    matched_df = pd.DataFrame(matches)
    bank_only = bank[~bank["matched"]][["source_row", "date", "description", "amount", "reference"]]
    ledger_only = ledger[~ledger["matched"]][["source_row", "date", "description", "amount", "reference"]]

    # --- Summary sheet with control-total check ---
    total_bank = bank["amount"].sum()
    total_ledger = ledger["amount"].sum()
    matched_bank_total = bank[bank["matched"]]["amount"].sum()

    summary = pd.DataFrame(
        {
            "Metric": [
                "Bank rows (total)", "Ledger rows (total)",
                "Matched pairs (total)", "Full (exact) matches", "Fuzzy/LLM matches",
                "Bank rows unmatched", "Ledger rows unmatched",
                "Match rate (bank)",
                "Sum of all bank amounts", "Sum of all ledger amounts",
                "Sum of matched bank amounts", "Sum of unmatched bank amounts",
                "Control check (matched + unmatched bank == total bank)",
            ],
            "Value": [
                len(bank), len(ledger),
                len(matched_df), len(exact_matches), len(fuzzy_matches),
                len(bank_only), len(ledger_only),
                f"{(bank['matched'].sum() / max(len(bank), 1)) * 100:.1f}%",
                round(total_bank, 2), round(total_ledger, 2),
                round(matched_bank_total, 2), round(bank[~bank["matched"]]["amount"].sum(), 2),
                "PASS" if abs((matched_bank_total + bank[~bank["matched"]]["amount"].sum()) - total_bank) < 0.01 else "FAIL",
            ],
        }
    )
    _write_df_sheet(wb, "Summary", summary)

    # --- Matches, split by how confidently they were determined ---
    _write_df_sheet(wb, "Full_Matches", _prepare_match_sheet_df(exact_matches), row_fill=GOOD_FILL)
    _write_df_sheet(wb, "Fuzzy_Matches", _prepare_match_sheet_df(fuzzy_matches), row_fill=WARN_FILL)

    # --- Needs_Review: LLM was consulted but did NOT auto-accept -------
    # (low confidence, or no plausible candidate at all). This is the
    # sheet the module docstring always promised but the code never
    # actually built -- LLM_Audit_Log logs EVERY call regardless of
    # outcome, which is too noisy for "what do I need to personally
    # decide on" -- this sheet is just that subset.
    review_rows = [
        r for r in llm_log
        if r["result"].get("confidence", 0) < llm_confidence_threshold
        or r["result"].get("match_ledger_row") is None
    ]
    if review_rows:
        review_df = pd.DataFrame(
            [
                {
                    "bank_row": r["bank_row"]["row"],
                    "bank_date": r["bank_row"]["date"],
                    "bank_description": r["bank_row"]["description"],
                    "bank_amount": r["bank_row"]["amount"],
                    "candidates_considered": len(r["candidates"]),
                    "llm_suggested_ledger_row": r["result"].get("match_ledger_row"),
                    "llm_confidence": r["result"].get("confidence"),
                    "llm_rationale": r["result"].get("rationale"),
                }
                for r in review_rows
            ]
        )
    else:
        review_df = pd.DataFrame()
    _write_df_sheet(wb, "Needs_Review", review_df, row_fill=WARN_FILL)

    _write_df_sheet(wb, "Bank_Only_Unmatched", bank_only, row_fill=WARN_FILL)
    _write_df_sheet(wb, "Ledger_Only_Unmatched", ledger_only, row_fill=WARN_FILL)

    if llm_log:
        log_rows = []
        for entry in llm_log:
            log_rows.append(
                {
                    "bank_row": entry["bank_row"]["row"],
                    "bank_description": entry["bank_row"]["description"],
                    "bank_amount": entry["bank_row"]["amount"],
                    "candidates_considered": len(entry["candidates"]),
                    "llm_decision_row": entry["result"].get("match_ledger_row"),
                    "llm_confidence": entry["result"].get("confidence"),
                    "llm_rationale": entry["result"].get("rationale"),
                }
            )
        _write_df_sheet(wb, "LLM_Audit_Log", pd.DataFrame(log_rows))

    wb.save(path)